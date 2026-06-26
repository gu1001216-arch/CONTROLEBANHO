"""
Controle de Produtividade Banho — Pintura Eletrostática
Fluxo: Preparação (1 ou 2 operadores) -> Preencher OPs -> Fila do banho -> Banho -> Concluído

Recursos:
- Grade de 19 cestos com cadeado
- Pausar/retomar o tempo de preparação (café, ginástica) — desconta do total
- Parar o tempo e só depois preencher os dados
- Múltiplas OPs por cesto (botão Adicionar OP)
- 1 ou 2 operadores por cesto (definido ao iniciar)
- Lista mestra do SAP carregada do arquivo (lista_mestra.xlsx, .csv ou .txt)
  direto na memória RAM ao iniciar — sem banco de dados para isso.
  Para atualizar: substitua o arquivo no repositório e faça redeploy.
- Dashboards (admin e público) + export Excel pré-banho e banho

Banco: PostgreSQL (Railway) apenas para cards e usuários. Local sem DATABASE_URL -> SQLite.
"""
import os
import io
import csv
import json
import threading
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Float, Text, inspect, text
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'troque-esta-chave-em-producao')

# ─────────────────────────────────────────────────────────────────────────────
# Lista mestra em memória (carregada do arquivo — sem banco de dados)
# ─────────────────────────────────────────────────────────────────────────────
_lista_lock = threading.Lock()
_lista_por_ordem = {}
_lista_por_material = {}
_lista_status = {'carregada': False, 'total': 0, 'erro': None}

LISTA_MESTRA_ARQUIVOS = [
    'lista_mestra.xlsx',
    'lista_mestra.csv',
    'lista_mestra.txt',
    'exemplo_lista_mestra_sap.txt',
]


def _norm_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _achar_arquivo_mestre():
    base = os.path.dirname(os.path.abspath(__file__))
    for nome in LISTA_MESTRA_ARQUIVOS:
        caminho = os.path.join(base, nome)
        if os.path.isfile(caminho):
            return caminho
    return None


def _parsear_linhas_mestre(linhas):
    achado = _achar_colunas(linhas)
    if achado:
        cab_idx, col = achado
        i_ordem = col.get('ordem', 0)
        i_mat   = col.get('material', 1)
        i_texto = col.get('texto')
        i_qtd   = col.get('qtd')
        inicio  = cab_idx + 1
    else:
        i_ordem, i_mat, i_texto, i_qtd = 0, 2, 3, 4
        inicio = 0

    def val(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    por_ordem    = {}
    por_material = {}
    for row in linhas[inicio:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        ordem = _norm_ordem(row[i_ordem]) if i_ordem < len(row) and row[i_ordem] is not None else ''
        if not ordem or not ordem.replace('.', '').isdigit():
            continue
        material   = val(row, i_mat)
        texto      = val(row, i_texto)
        q          = val(row, i_qtd)
        try:
            qtd = int(float(q)) if q else 0
        except (ValueError, TypeError):
            qtd = 0
        item = {'ordem': ordem, 'material': material, 'texto_breve': texto, 'quantidade': qtd}
        por_ordem[ordem] = item
        if material and material not in por_material:
            por_material[material] = item
    return por_ordem, por_material


# ── Área/Peso por código SAP (mesma abordagem: arquivo em memória) ──────────
_areapeso_por_sap = {}
_areapeso_status = {'carregada': False, 'total': 0, 'erro': None}
AREA_PESO_ARQUIVOS = ['area_peso.xlsx', 'area_peso.csv', 'area_peso.txt']


def _achar_arquivo_areapeso():
    base = os.path.dirname(os.path.abspath(__file__))
    for nome in AREA_PESO_ARQUIVOS:
        caminho = os.path.join(base, nome)
        if os.path.isfile(caminho):
            return caminho
    return None


def carregar_area_peso():
    """Carrega area_peso (Codigo SAP -> área superfície e peso unitários)."""
    global _areapeso_por_sap, _areapeso_status
    caminho = _achar_arquivo_areapeso()
    if not caminho:
        _areapeso_status = {'carregada': False, 'total': 0, 'erro': 'Arquivo area_peso não encontrado.'}
        print('[area_peso] AVISO: nenhum arquivo encontrado.')
        return
    try:
        linhas = []
        nome = caminho.lower()
        if nome.endswith('.csv') or nome.endswith('.txt'):
            with open(caminho, encoding='utf-8-sig', errors='replace') as f:
                raw = f.read()
            sep = '\t' if raw.count('\t') > raw.count(';') and raw.count('\t') > raw.count(',') \
                else (';' if raw.count(';') > raw.count(',') else ',')
            linhas = list(csv.reader(io.StringIO(raw), delimiter=sep))
        else:
            from openpyxl import load_workbook as _lw
            wb = _lw(caminho, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                linhas.append(list(row))

        # detecta colunas pelo nome do cabeçalho
        def norm(s):
            return str(s).strip().lower() if s is not None else ''
        i_sap = i_area = i_peso = None
        inicio = 0
        for i, row in enumerate(linhas[:10]):
            if not row:
                continue
            nomes = [norm(c) for c in row]
            for j, n in enumerate(nomes):
                if 'codigo sap' in n or 'código sap' in n:
                    i_sap = j
                elif 'area' in n or 'área' in n:
                    i_area = j
                elif 'peso' in n:
                    i_peso = j
            if i_sap is not None:
                inicio = i + 1
                break
        if i_sap is None:  # fallback p/ a ordem do arquivo: Codigo, Area, Peso, Codigo Sap
            i_sap, i_area, i_peso, inicio = 3, 1, 2, 1

        mapa = {}
        for row in linhas[inicio:]:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            sap = _norm_str(row[i_sap]) if i_sap is not None and i_sap < len(row) else ''
            if not sap:
                continue
            def num(idx):
                try:
                    return float(row[idx]) if idx is not None and idx < len(row) and row[idx] not in (None, '') else 0.0
                except (ValueError, TypeError):
                    return 0.0
            area_mm2 = num(i_area)
            peso_kg = num(i_peso)
            mapa[sap] = {'area_m2': area_mm2 / 1_000_000.0,  # mm² -> m²
                         'peso_kg': peso_kg}
        _areapeso_por_sap = mapa
        _areapeso_status = {'carregada': True, 'total': len(mapa), 'erro': None}
        print(f'[area_peso] Carregada: {len(mapa)} códigos de "{os.path.basename(caminho)}".')
    except Exception as e:
        _areapeso_status = {'carregada': False, 'total': 0, 'erro': str(e)}
        print(f'[area_peso] ERRO ao carregar: {e}')


def _area_peso_do_codigo(material):
    """Retorna (area_m2, peso_kg) unitários para um código SAP, ou (0,0)."""
    d = _areapeso_por_sap.get(_norm_str(material))
    if d:
        return d['area_m2'], d['peso_kg']
    return 0.0, 0.0


def carregar_lista_mestre():
    global _lista_por_ordem, _lista_por_material, _lista_status
    caminho = _achar_arquivo_mestre()
    if not caminho:
        with _lista_lock:
            _lista_status = {'carregada': False, 'total': 0,
                             'erro': 'Arquivo lista_mestra.xlsx/.csv/.txt não encontrado na raiz do projeto.'}
        print('[lista_mestra] AVISO: nenhum arquivo encontrado.')
        return

    try:
        nome = caminho.lower()
        linhas = []
        if nome.endswith('.csv') or nome.endswith('.txt'):
            with open(caminho, encoding='utf-8-sig', errors='replace') as f:
                raw = f.read()
            sep = '\t' if raw.count('\t') > raw.count(';') and raw.count('\t') > raw.count(',') \
                else (';' if raw.count(';') > raw.count(',') else ',')
            linhas = list(csv.reader(io.StringIO(raw), delimiter=sep))
        else:
            from openpyxl import load_workbook as _lw
            wb = _lw(caminho, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                linhas.append(list(row))

        por_ordem, por_material = _parsear_linhas_mestre(linhas)
        with _lista_lock:
            _lista_por_ordem    = por_ordem
            _lista_por_material = por_material
            _lista_status = {'carregada': True, 'total': len(por_ordem), 'erro': None}
        print(f'[lista_mestra] Carregada: {len(por_ordem)} ordens de "{os.path.basename(caminho)}".')
    except Exception as e:
        with _lista_lock:
            _lista_status = {'carregada': False, 'total': 0, 'erro': str(e)}
        print(f'[lista_mestra] ERRO ao carregar: {e}')

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

_EM_PRODUCAO = bool(os.environ.get('RAILWAY_ENVIRONMENT') or
                    os.environ.get('RAILWAY_PROJECT_ID') or os.environ.get('PORT'))

if not DATABASE_URL:
    if _EM_PRODUCAO:
        print('=' * 70)
        print('ERRO CRÍTICO: DATABASE_URL não definida! Sem ela, o app usa SQLite')
        print('temporário e os dados se perdem a cada deploy. No Railway, serviço')
        print('"web", em Variables, adicione: DATABASE_URL = ${{Postgres.DATABASE_URL}}')
        print('=' * 70)
    DATABASE_URL = 'sqlite:///dados_local.db'

engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=280)
Session = scoped_session(sessionmaker(bind=engine))
Base = declarative_base()

TOTAL_CESTOS = 20

# Senha de acesso ao painel público (gerência)
SENHA_PAINEL = os.environ.get('SENHA_PAINEL', 'Decio2026@')

PROCESSOS = [
    "AÇO SEM OXIDAÇÃO", "AÇO COM OXIDAÇÃO", "ALUMÍNIO",
    "MINIMIZADO SEM OXIDAÇÃO", "MINIMIZADO COM OXIDAÇÃO", "INOX",
]

ST_PREPARANDO = 'PREPARANDO'
ST_PREENCHER = 'PREENCHER'
ST_FILA_BANHO = 'FILA_BANHO'
ST_EM_BANHO = 'EM_BANHO'
ST_CONCLUIDO = 'CONCLUIDO'
ESTADOS_ATIVOS = (ST_PREPARANDO, ST_PREENCHER, ST_FILA_BANHO, ST_EM_BANHO)


# ─────────────────────────────────────────────────────────────────────────────
# Modelos
# ─────────────────────────────────────────────────────────────────────────────
class Usuario(Base):
    __tablename__ = 'usuarios'
    id = Column(Integer, primary_key=True)
    login = Column(String(50), unique=True, nullable=False)
    nome = Column(String(120), nullable=False)
    senha_hash = Column(String(255), nullable=False)
    perfil = Column(String(20), nullable=False)

    def to_dict(self):
        return {'id': self.id, 'login': self.login, 'nome': self.nome, 'perfil': self.perfil}


def _pausas_resumo(pausas_json):
    try:
        mapa = json.loads(pausas_json) if pausas_json else {}
    except (ValueError, TypeError):
        mapa = {}
    lista = [{'motivo': m, 'minutos': round(s / 60, 1)} for m, s in mapa.items()]
    texto = '; '.join(f"{x['motivo']}: {x['minutos']} min" for x in lista)
    total_seg = sum(mapa.values())
    return {'lista': lista, 'texto': texto, 'total_seg': total_seg}


class Card(Base):
    __tablename__ = 'cards'
    id = Column(Integer, primary_key=True)
    estado = Column(String(20), nullable=False, index=True)

    numero_cesto = Column(Integer, nullable=False)
    processo = Column(String(60), default='')
    tipo = Column(String(20), default='Normal')

    ordem = Column(String(60), default='')
    material = Column(String(60), default='')
    texto_breve = Column(String(255), default='')
    quantidade = Column(Integer, default=0)
    itens_json = Column(Text, default='')
    observacao = Column(Text, default='')

    operador_prep = Column(String(120), default='')
    operador_prep2 = Column(String(120), default='')
    n_operadores = Column(Integer, default=1)
    # operador que INICIOU o banho
    operador_banho_inicio = Column(String(120), default='')
    # operador que FINALIZOU o banho
    operador_banho_fim = Column(String(120), default='')
    # campo legado (mantido para compatibilidade)
    operador_banho = Column(String(120), default='')

    prep_inicio = Column(DateTime)
    prep_fim = Column(DateTime)
    prep_minutos = Column(Float, default=0)

    pausado = Column(Integer, default=0)
    pausa_inicio = Column(DateTime)
    pausa_motivo = Column(String(60), default='')
    pausa_acumulada_seg = Column(Integer, default=0)
    pausas_json = Column(Text, default='')

    banho_inicio = Column(DateTime)
    banho_fim = Column(DateTime)
    banho_minutos = Column(Float, default=0)
    obs_banho = Column(Text, default='')

    criado_em = Column(DateTime, default=datetime.utcnow)

    def to_dict(self):
        def fmt(dt):
            return (dt - timedelta(hours=3)).strftime('%d/%m/%Y %H:%M:%S') if dt else ''

        def fmt_data(dt):
            return (dt - timedelta(hours=3)).strftime('%d/%m/%Y') if dt else ''

        def fmt_hora(dt):
            return (dt - timedelta(hours=3)).strftime('%H:%M:%S') if dt else ''

        def iso(dt):
            return dt.isoformat() + 'Z' if dt else ''
        try:
            itens = json.loads(self.itens_json) if self.itens_json else []
        except (ValueError, TypeError):
            itens = []
        if not itens and self.ordem:
            itens = [{'ordem': self.ordem, 'material': self.material,
                      'texto_breve': self.texto_breve, 'quantidade': self.quantidade}]
        qtd_total = sum(int(i.get('quantidade') or 0) for i in itens) if itens else (self.quantidade or 0)
        # peso e área totais = soma de (unitário do código SAP × quantidade do item)
        peso_total = 0.0
        area_total = 0.0
        for it in itens:
            a_unit, p_unit = _area_peso_do_codigo(it.get('material', ''))
            q = int(it.get('quantidade') or 0)
            area_total += a_unit * q
            peso_total += p_unit * q
        pausas = _pausas_resumo(self.pausas_json)
        total_pausa_min = round(pausas['total_seg'] / 60, 1)
        # tempo de espera = início do banho − fim da preparação (parado na fila)
        espera_min = 0.0
        if self.banho_inicio and self.prep_fim:
            espera_min = round(max(0, (self.banho_inicio - self.prep_fim).total_seconds()) / 60, 1)
        return {
            'id': self.id, 'estado': self.estado,
            'numero_cesto': self.numero_cesto,
            'processo': self.processo, 'tipo': self.tipo,
            'ordem': self.ordem, 'material': self.material,
            'texto_breve': self.texto_breve, 'quantidade': self.quantidade,
            'itens': itens, 'qtd_total': qtd_total, 'n_itens': len(itens),
            'peso_total': round(peso_total, 2), 'area_total': round(area_total, 3),
            'observacao': self.observacao or '',
            'obs_banho': self.obs_banho or '',
            'operador_prep': self.operador_prep, 'operador_prep2': self.operador_prep2 or '',
            'n_operadores': self.n_operadores or 1,
            'operador_banho': self.operador_banho or '',
            'operador_banho_inicio': self.operador_banho_inicio or self.operador_banho or '',
            'operador_banho_fim': self.operador_banho_fim or '',
            'prep_inicio': fmt(self.prep_inicio), 'prep_fim': fmt(self.prep_fim),
            'prep_inicio_data': fmt_data(self.prep_inicio), 'prep_inicio_hora': fmt_hora(self.prep_inicio),
            'prep_fim_data': fmt_data(self.prep_fim), 'prep_fim_hora': fmt_hora(self.prep_fim),
            'prep_minutos': round(self.prep_minutos or 0, 1),
            'total_pausa_min': total_pausa_min,
            'espera_min': espera_min,
            'banho_inicio': fmt(self.banho_inicio), 'banho_fim': fmt(self.banho_fim),
            'banho_inicio_data': fmt_data(self.banho_inicio),
            'banho_inicio_hora': fmt_hora(self.banho_inicio),
            'banho_fim_data': fmt_data(self.banho_fim),
            'banho_fim_hora': fmt_hora(self.banho_fim),
            'banho_minutos': round(self.banho_minutos or 0, 1),
            'prep_inicio_iso': iso(self.prep_inicio),
            'prep_fim_iso': iso(self.prep_fim),
            'banho_inicio_iso': iso(self.banho_inicio),
            'pausado': bool(self.pausado),
            'pausa_inicio_iso': iso(self.pausa_inicio),
            'pausa_motivo': self.pausa_motivo or '',
            'pausa_acumulada_seg': self.pausa_acumulada_seg or 0,
            'pausas': pausas,
            'data_ref': (self.banho_fim - timedelta(hours=3)).strftime('%Y-%m-%d') if self.banho_fim else '',
        }


# ─────────────────────────────────────────────────────────────────────────────
# Init + migração + seed
# ─────────────────────────────────────────────────────────────────────────────
def _migrar_colunas():
    insp = inspect(engine)
    if 'cards' not in insp.get_table_names():
        return
    existentes = {c['name'] for c in insp.get_columns('cards')}
    novas = {
        'itens_json': 'TEXT', 'operador_prep2': "VARCHAR(120) DEFAULT ''",
        'n_operadores': 'INTEGER DEFAULT 1', 'pausado': 'INTEGER DEFAULT 0',
        'pausa_inicio': 'TIMESTAMP NULL', 'pausa_acumulada_seg': 'INTEGER DEFAULT 0',
        'pausa_motivo': "VARCHAR(60) DEFAULT ''", 'pausas_json': 'TEXT',
        'obs_banho': 'TEXT',
        'operador_banho_inicio': "VARCHAR(120) DEFAULT ''",
        'operador_banho_fim': "VARCHAR(120) DEFAULT ''",
    }
    with engine.begin() as conn:
        for col, tipo in novas.items():
            if col not in existentes:
                try:
                    conn.execute(text(f'ALTER TABLE cards ADD COLUMN {col} {tipo}'))
                except Exception:
                    pass


def init_db():
    Base.metadata.create_all(engine)
    _migrar_colunas()
    carregar_lista_mestre()
    carregar_area_peso()
    db = Session()
    try:
        if db.query(Usuario).count() == 0:
            seed = [('admin', 'Administrador', 'admin123', 'admin'),
                    ('banho', 'Operador de Banho', 'banho123', 'banho')]
            for i in range(1, 7):
                seed.append((f'op{i}', f'Operador {i}', 'op1234', 'prep'))
            for login, nome, senha, perfil in seed:
                db.add(Usuario(login=login, nome=nome,
                               senha_hash=generate_password_hash(senha), perfil=perfil))
            db.commit()
    finally:
        db.close()


def login_required(*perfis):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if 'usuario' not in session:
                return redirect(url_for('login'))
            if perfis and session.get('perfil') not in perfis and session.get('perfil') != 'admin':
                return redirect(url_for('login'))
            return f(*a, **kw)
        return wrapper
    return deco


def _norm_ordem(v):
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    # Remove 4 dígitos do início e 4 dígitos do fim se o resultado tiver 8 dígitos (ou mais)
    # Ex: código de barras com 16 dígitos: XXXX[8digitos]XXXX -> extrai os 8 do meio
    s_digits = ''.join(c for c in s if c.isdigit())
    if len(s_digits) > 8:
        # Remove 4 prefixos e 4 sufixos numéricos
        s_digits = s_digits[4:-4]
    return s_digits if s_digits else s


# ─────────────────────────────────────────────────────────────────────────────
# Páginas
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        login_u = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '')
        db = Session()
        try:
            u = db.query(Usuario).filter_by(login=login_u).first()
            if u and check_password_hash(u.senha_hash, senha):
                session['usuario'] = u.login
                session['nome'] = u.nome
                session['perfil'] = u.perfil
                destino = {'admin': 'dashboard', 'banho': 'tela_banho',
                           'prep': 'tela_prep'}.get(u.perfil, 'login')
                return redirect(url_for(destino))
            erro = 'Usuário ou senha incorretos.'
        finally:
            db.close()
    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


def _lista_operadores_prep():
    db = Session()
    try:
        return [u.nome for u in db.query(Usuario)
                .filter(Usuario.perfil.in_(('prep', 'banho', 'admin')))
                .order_by(Usuario.nome).all()]
    finally:
        db.close()


@app.route('/preparacao')
@login_required('prep', 'banho')
def tela_prep():
    return render_template('prep.html', nome=session.get('nome'),
                           perfil=session.get('perfil'), processos=PROCESSOS)


@app.route('/banho')
@login_required('banho')
def tela_banho():
    return render_template('banho.html', nome=session.get('nome'), perfil=session.get('perfil'))


@app.route('/dashboard')
@login_required('admin')
def dashboard():
    return render_template('dashboard.html', nome=session.get('nome'), processos=PROCESSOS)


@app.route('/painel', methods=['GET', 'POST'])
def painel_publico():
    if request.method == 'POST':
        if request.form.get('senha', '') == SENHA_PAINEL:
            session['painel_ok'] = True
            return redirect(url_for('painel_publico'))
        return render_template('painel_senha.html', erro='Senha incorreta.')
    if not session.get('painel_ok'):
        return render_template('painel_senha.html', erro=None)
    return render_template('painel.html', processos=PROCESSOS)


@app.route('/admin/usuarios', methods=['GET', 'POST'])
@login_required('admin')
def admin_usuarios():
    db = Session()
    msg = None
    try:
        if request.method == 'POST':
            acao = request.form.get('acao')
            if acao == 'adicionar':
                nu = request.form.get('novo_usuario', '').strip()
                nn = request.form.get('novo_nome', '').strip()
                ns = request.form.get('nova_senha', '')
                npf = request.form.get('novo_perfil', 'prep')
                if nu and nn and ns and not db.query(Usuario).filter_by(login=nu).first():
                    db.add(Usuario(login=nu, nome=nn,
                                   senha_hash=generate_password_hash(ns), perfil=npf))
                    db.commit()
                    msg = f'Usuário {nn} adicionado.'
                elif db.query(Usuario).filter_by(login=nu).first():
                    msg = 'Esse login já existe.'
            elif acao == 'remover':
                u = db.query(Usuario).filter_by(login=request.form.get('usuario_remover')).first()
                if u and u.login != 'admin':
                    db.delete(u)
                    db.commit()
                    msg = 'Usuário removido.'
            elif acao == 'senha':
                u = db.query(Usuario).filter_by(login=request.form.get('usuario_senha')).first()
                nova = request.form.get('senha_nova', '')
                if u and nova:
                    u.senha_hash = generate_password_hash(nova)
                    db.commit()
                    msg = f'Senha de {u.nome} atualizada.'
        usuarios = [u.to_dict() for u in db.query(Usuario).order_by(Usuario.id).all()]
        return render_template('usuarios.html', usuarios=usuarios, nome=session.get('nome'), msg=msg)
    finally:
        db.close()


@app.route('/admin/mestre', methods=['GET', 'POST'])
@login_required('admin')
def admin_mestre():
    msg = None
    if request.method == 'POST' and request.form.get('acao') == 'recarregar':
        carregar_lista_mestre()
        carregar_area_peso()
        st = _lista_status
        if st['carregada']:
            msg = f'Lista recarregada com sucesso: {st["total"]} ordens na memória.'
        else:
            msg = f'Erro ao recarregar: {st["erro"]}'

    with _lista_lock:
        st     = dict(_lista_status)
        amostra = list(_lista_por_ordem.values())[:25]

    caminho = _achar_arquivo_mestre()
    arquivo_info = os.path.basename(caminho) if caminho else 'Nenhum arquivo encontrado'
    ap = _areapeso_status
    cap = _achar_arquivo_areapeso()
    areapeso_info = os.path.basename(cap) if cap else 'Nenhum arquivo encontrado'
    return render_template('mestre.html', nome=session.get('nome'),
                           msg=msg, total_itens=st['total'],
                           amostra=amostra, status=st,
                           arquivo_info=arquivo_info,
                           areapeso_status=ap, areapeso_info=areapeso_info)


@app.route('/api/admin/testar_op/<path:ordem>')
@login_required('admin')
def api_admin_testar_op(ordem):
    o = _norm_ordem(ordem)
    with _lista_lock:
        item = _lista_por_ordem.get(o)
    if item:
        return jsonify({'encontrado': True, **item})
    return jsonify({'encontrado': False, 'ordem': o})


def _achar_colunas(linhas):
    def norm(s):
        return str(s).strip().lower() if s is not None else ''
    for i, row in enumerate(linhas[:10]):
        if not row:
            continue
        nomes = [norm(c) for c in row]
        idx = {}
        for j, nome in enumerate(nomes):
            if nome == 'ordem' and 'ordem' not in idx:
                idx['ordem'] = j
            elif nome == 'material' and 'material' not in idx:
                idx['material'] = j
            elif 'texto breve' in nome and 'texto' not in idx:
                idx['texto'] = j
            elif ('quantidade da ordem' in nome or nome == 'quantidade total'
                  or nome == 'quantidade') and 'qtd' not in idx:
                idx['qtd'] = j
        if 'ordem' in idx and 'material' in idx:
            return i, idx
    return None


# ─────────────────────────────────────────────────────────────────────────────
# APIs — grade e fluxo
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/cestos')
@login_required('prep', 'banho')
def api_cestos():
    db = Session()
    try:
        ativos = db.query(Card).filter(Card.estado.in_(ESTADOS_ATIVOS)).all()
        mapa = {c.numero_cesto: c for c in ativos}
        grade = [{'numero': n, 'ocupado': n in mapa,
                  'card': mapa[n].to_dict() if n in mapa else None}
                 for n in range(1, TOTAL_CESTOS + 1)]
        return jsonify(grade)
    finally:
        db.close()


@app.route('/api/agora')
def api_agora():
    """Hora atual do servidor (UTC ISO) — p/ sincronizar cronômetros e começar do 0:00."""
    return jsonify({'agora_iso': datetime.utcnow().isoformat() + 'Z'})


def _serial_val(v):
    if isinstance(v, datetime):
        return v.isoformat() + 'Z'
    return v


@app.route('/api/admin/db_status')
@login_required('admin')
def api_db_status():
    tipo = 'postgresql' if DATABASE_URL.startswith('postgresql') else 'sqlite'
    db = Session()
    try:
        n_cards = db.query(Card).count()
        n_users = db.query(Usuario).count()
    except Exception:
        n_cards = n_users = -1
    finally:
        db.close()
    return jsonify({'tipo': tipo, 'seguro': tipo == 'postgresql',
                    'cards': n_cards, 'usuarios': n_users,
                    'mestre': _lista_status.get('total', 0)})


@app.route('/api/admin/backup')
@login_required('admin')
def api_admin_backup():
    db = Session()
    try:
        def full(obj, model):
            return {c.name: _serial_val(getattr(obj, c.name)) for c in model.__table__.columns}
        dados = {
            'versao': 1, 'gerado_em': datetime.utcnow().isoformat() + 'Z',
            'cards': [full(c, Card) for c in db.query(Card).all()],
            'usuarios': [full(u, Usuario) for u in db.query(Usuario).all()],
        }
        buf = io.BytesIO(json.dumps(dados, ensure_ascii=False, indent=2).encode('utf-8'))
        buf.seek(0)
        stamp = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(buf, as_attachment=True,
                         download_name=f'backup_banho_{stamp}.json',
                         mimetype='application/json')
    finally:
        db.close()


@app.route('/api/admin/restaurar', methods=['POST'])
@login_required('admin')
def api_admin_restaurar():
    f = request.files.get('arquivo')
    if not f or not f.filename:
        return jsonify({'sucesso': False, 'erro': 'Envie o arquivo de backup.'}), 400
    try:
        dados = json.loads(f.stream.read().decode('utf-8-sig', errors='replace'))
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': f'Arquivo inválido: {e}'}), 400
    db = Session()
    rc = ru = 0
    try:
        def set_cols(obj, model, src):
            for col in model.__table__.columns:
                if col.name in src:
                    val = src[col.name]
                    if isinstance(col.type, DateTime) and val:
                        try:
                            val = datetime.fromisoformat(str(val).replace('Z', ''))
                        except (ValueError, TypeError):
                            val = None
                    setattr(obj, col.name, val)
        logins = {u.login for u in db.query(Usuario).all()}
        for u in dados.get('usuarios', []):
            if u.get('login') and u['login'] not in logins:
                novo = Usuario(login=u['login'], nome=u.get('nome', ''),
                               senha_hash=u.get('senha_hash', ''), perfil=u.get('perfil', 'prep'))
                db.add(novo); ru += 1
        ids = {c.id for c in db.query(Card.id).all()}
        for cd in dados.get('cards', []):
            if cd.get('id') and cd['id'] not in ids:
                novo = Card(); set_cols(novo, Card, cd); db.add(novo); rc += 1
        db.commit()
        return jsonify({'sucesso': True, 'cards': rc, 'usuarios': ru})
    except Exception as e:
        db.rollback()
        return jsonify({'sucesso': False, 'erro': str(e)}), 500
    finally:
        db.close()


@app.route('/api/cesto/mudar_numero', methods=['POST'])
@login_required('prep', 'banho')
def api_cesto_mudar_numero():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card:
            return jsonify({'sucesso': False, 'erro': 'Cesto não encontrado.'}), 404
        try:
            novo = int(d.get('numero_cesto'))
        except (ValueError, TypeError):
            return jsonify({'sucesso': False, 'erro': 'Número inválido.'}), 400
        if not (1 <= novo <= TOTAL_CESTOS):
            return jsonify({'sucesso': False, 'erro': f'O número deve ser entre 1 e {TOTAL_CESTOS}.'}), 400
        ocupado = db.query(Card).filter(Card.numero_cesto == novo,
                                        Card.estado.in_(ESTADOS_ATIVOS),
                                        Card.id != card.id).first()
        if ocupado:
            return jsonify({'sucesso': False, 'erro': f'O cesto {novo} já está em uso.'}), 400
        card.numero_cesto = novo
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/cesto/cancelar', methods=['POST'])
@login_required('prep', 'banho')
def api_cesto_cancelar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card:
            return jsonify({'sucesso': False, 'erro': 'Cesto não encontrado.'}), 404
        if card.estado == ST_CONCLUIDO:
            return jsonify({'sucesso': False, 'erro': 'Não é possível cancelar um cesto concluído.'}), 400
        db.delete(card)
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/admin/lista_status')
@login_required('admin')
def api_lista_status():
    with _lista_lock:
        st = dict(_lista_status)
    st['arquivo'] = os.path.basename(_achar_arquivo_mestre() or '') or 'não encontrado'
    return jsonify(st)


@app.route('/api/buscar_ordem/<path:ordem>')
@login_required('prep', 'banho')
def api_buscar_ordem(ordem):
    o = _norm_ordem(ordem)
    with _lista_lock:
        item = _lista_por_ordem.get(o)
    if item:
        return jsonify({'encontrado': True, **item})
    return jsonify({'encontrado': False, 'ordem': o})


@app.route('/api/buscar_codigo/<path:codigo>')
@login_required('prep', 'banho')
def api_buscar_codigo(codigo):
    cod = _norm_ordem(codigo)
    with _lista_lock:
        item = _lista_por_material.get(cod)
    if item:
        return jsonify({'encontrado': True, 'material': item['material'],
                        'texto_breve': item['texto_breve'], 'quantidade': item['quantidade']})
    return jsonify({'encontrado': False, 'material': cod})


@app.route('/api/prep/iniciar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_iniciar():
    d = request.json or {}
    try:
        numero = int(d.get('numero_cesto'))
    except (ValueError, TypeError):
        return jsonify({'sucesso': False, 'erro': 'Cesto inválido.'}), 400
    if not (1 <= numero <= TOTAL_CESTOS):
        return jsonify({'sucesso': False, 'erro': 'Cesto fora do intervalo.'}), 400
    db = Session()
    try:
        if db.query(Card).filter(Card.numero_cesto == numero,
                                 Card.estado.in_(ESTADOS_ATIVOS)).first():
            return jsonify({'sucesso': False, 'erro': f'Cesto {numero} já está em uso.'}), 400
        try:
            n_op = int(d.get('n_operadores', 1))
            n_op = n_op if n_op in (1, 2, 3) else 1
        except (ValueError, TypeError):
            n_op = 1
        card = Card(estado=ST_PREPARANDO, numero_cesto=numero,
                    operador_prep=session.get('nome', ''),
                    operador_prep2='', n_operadores=n_op,
                    prep_inicio=datetime.utcnow())
        db.add(card)
        db.commit()
        return jsonify({'sucesso': True, 'id': card.id})
    finally:
        db.close()


@app.route('/api/prep/pausar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_pausar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_PREPARANDO:
            return jsonify({'sucesso': False, 'erro': 'Cesto não está em preparação.'}), 404
        agora = datetime.utcnow()
        if card.pausado:
            if card.pausa_inicio:
                dur = int((agora - card.pausa_inicio).total_seconds())
                card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + dur
                try:
                    mapa = json.loads(card.pausas_json) if card.pausas_json else {}
                except (ValueError, TypeError):
                    mapa = {}
                motivo = card.pausa_motivo or 'Outros'
                mapa[motivo] = mapa.get(motivo, 0) + dur
                card.pausas_json = json.dumps(mapa, ensure_ascii=False)
            card.pausado = 0
            card.pausa_inicio = None
            card.pausa_motivo = ''
        else:
            motivo = (d.get('motivo') or '').strip()
            if not motivo:
                return jsonify({'sucesso': False, 'erro': 'Informe o motivo da pausa.'}), 400
            card.pausado = 1
            card.pausa_inicio = agora
            card.pausa_motivo = motivo[:60]
        db.commit()
        return jsonify({'sucesso': True, 'pausado': bool(card.pausado)})
    finally:
        db.close()


@app.route('/api/prep/parar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_parar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_PREPARANDO:
            return jsonify({'sucesso': False, 'erro': 'Cesto não está em preparação.'}), 404
        agora = datetime.utcnow()
        if card.pausado and card.pausa_inicio:
            dur = int((agora - card.pausa_inicio).total_seconds())
            card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + dur
            try:
                mapa = json.loads(card.pausas_json) if card.pausas_json else {}
            except (ValueError, TypeError):
                mapa = {}
            motivo = card.pausa_motivo or 'Outros'
            mapa[motivo] = mapa.get(motivo, 0) + dur
            card.pausas_json = json.dumps(mapa, ensure_ascii=False)
            card.pausado = 0
            card.pausa_inicio = None
            card.pausa_motivo = ''
        card.prep_fim = agora
        bruto = (card.prep_fim - card.prep_inicio).total_seconds()
        card.prep_minutos = round(max(0, bruto - (card.pausa_acumulada_seg or 0)) / 60, 1)
        card.estado = ST_PREENCHER
        db.commit()
        return jsonify({'sucesso': True, 'prep_minutos': card.prep_minutos})
    finally:
        db.close()


def _salvar_itens(card, d):
    itens = d.get('itens')
    if itens is None:
        itens = [{'ordem': d.get('ordem', ''), 'material': d.get('material', ''),
                  'texto_breve': d.get('texto_breve', ''), 'quantidade': d.get('quantidade', 0)}]
    norm = []
    for it in itens:
        ordem = _norm_ordem(it.get('ordem', '')) if it.get('ordem') else ''
        if not ordem and not it.get('material'):
            continue
        try:
            q = int(it.get('quantidade') or 0)
        except (ValueError, TypeError):
            q = 0
        norm.append({'ordem': ordem, 'material': (it.get('material') or '').strip(),
                     'texto_breve': (it.get('texto_breve') or '').strip(), 'quantidade': q})
    card.itens_json = json.dumps(norm, ensure_ascii=False)
    if norm:
        card.ordem = norm[0]['ordem']
        card.material = norm[0]['material']
        card.texto_breve = norm[0]['texto_breve']
        card.quantidade = sum(i['quantidade'] for i in norm)


def _aplicar_meta(card, d):
    for campo in ('processo', 'tipo', 'observacao'):
        if campo in d:
            setattr(card, campo, (d.get(campo) or '').strip())


@app.route('/api/prep/finalizar', methods=['POST'])
@login_required('prep', 'banho')
def api_prep_finalizar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado not in (ST_PREENCHER, ST_PREPARANDO):
            return jsonify({'sucesso': False, 'erro': 'Card não encontrado.'}), 404
        if card.estado == ST_PREPARANDO:
            agora = datetime.utcnow()
            if card.pausado and card.pausa_inicio:
                card.pausa_acumulada_seg = (card.pausa_acumulada_seg or 0) + \
                    int((agora - card.pausa_inicio).total_seconds())
                card.pausado = 0
                card.pausa_inicio = None
            card.prep_fim = agora
            bruto = (card.prep_fim - card.prep_inicio).total_seconds()
            card.prep_minutos = round(max(0, bruto - (card.pausa_acumulada_seg or 0)) / 60, 1)
        # registra quem FINALIZOU a preparação deste cesto
        card.operador_prep = session.get('nome', '') or card.operador_prep
        _aplicar_meta(card, d)
        _salvar_itens(card, d)
        card.estado = ST_FILA_BANHO
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/card/editar', methods=['POST'])
@login_required('prep', 'banho')
def api_card_editar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card:
            return jsonify({'sucesso': False, 'erro': 'Card não encontrado.'}), 404
        _aplicar_meta(card, d)
        if 'itens' in d:
            _salvar_itens(card, d)
        if d.get('prep_minutos') not in (None, ''):
            try:
                card.prep_minutos = round(float(d.get('prep_minutos')), 1)
            except (ValueError, TypeError):
                pass
        if 'n_operadores' in d:
            try:
                n = int(d.get('n_operadores'))
                card.n_operadores = n if n in (1, 2, 3) else 1
            except (ValueError, TypeError):
                pass
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/banho/fila')
@login_required('banho')
def api_banho_fila():
    db = Session()
    try:
        fila = db.query(Card).filter_by(estado=ST_FILA_BANHO).order_by(Card.prep_fim).all()
        emb = db.query(Card).filter_by(estado=ST_EM_BANHO).order_by(Card.banho_inicio).all()
        return jsonify({'fila': [c.to_dict() for c in fila],
                        'em_banho': [c.to_dict() for c in emb]})
    finally:
        db.close()


@app.route('/api/banho/iniciar', methods=['POST'])
@login_required('banho')
def api_banho_iniciar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_FILA_BANHO:
            return jsonify({'sucesso': False, 'erro': 'Card não está na fila.'}), 404
        card.banho_inicio = datetime.utcnow()
        nome_op = session.get('nome', '')
        card.operador_banho_inicio = nome_op
        card.operador_banho = nome_op  # compatibilidade
        card.estado = ST_EM_BANHO
        db.commit()
        return jsonify({'sucesso': True})
    finally:
        db.close()


@app.route('/api/banho/finalizar', methods=['POST'])
@login_required('banho')
def api_banho_finalizar():
    d = request.json or {}
    db = Session()
    try:
        card = db.query(Card).get(int(d.get('id', 0)))
        if not card or card.estado != ST_EM_BANHO:
            return jsonify({'sucesso': False, 'erro': 'Card não está em banho.'}), 404
        card.banho_fim = datetime.utcnow()
        card.banho_minutos = round((card.banho_fim - card.banho_inicio).total_seconds() / 60, 1)
        card.obs_banho = (d.get('obs_banho') or '').strip()
        nome_op = session.get('nome', '')
        card.operador_banho_fim = nome_op
        card.estado = ST_CONCLUIDO
        db.commit()
        return jsonify({'sucesso': True, 'banho_minutos': card.banho_minutos})
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Dados dos dashboards
# ─────────────────────────────────────────────────────────────────────────────
def _coletar_dados(de=None, ate=None):
    db = Session()
    try:
        cards = db.query(Card).filter_by(estado=ST_CONCLUIDO).all()

        def dentro(c):
            if not c.banho_fim:
                return False
            dia = (c.banho_fim - timedelta(hours=3)).date()
            if de and dia < de:
                return False
            if ate and dia > ate:
                return False
            return True
        cards = [c for c in cards if dentro(c)]
        ativos = db.query(Card).filter(Card.estado.in_(ESTADOS_ATIVOS)).order_by(Card.id.desc()).all()
        normais = sum(1 for c in cards if c.tipo == 'Normal')
        retrab = sum(1 for c in cards if c.tipo == 'Retrabalho')
        # banho normal/retrabalho = só cestos com banho FINALIZADO (concluídos)
        banho_normal = normais
        banho_retrab = retrab
        tp = [c.prep_minutos for c in cards if c.prep_minutos]
        tb = [c.banho_minutos for c in cards if c.banho_minutos]
        # tempo de espera (fila) = início banho − fim preparação
        esperas = []
        for c in cards:
            if c.banho_inicio and c.prep_fim:
                esperas.append(max(0, (c.banho_inicio - c.prep_fim).total_seconds()) / 60)
        por_proc, por_dia = {}, {}
        peso_por_dia, area_por_dia = {}, {}
        peso_total_geral = 0.0
        area_total_geral = 0.0
        pecas_total_geral = 0
        for c in cards:
            p = c.processo or 'Sem processo'
            por_proc[p] = por_proc.get(p, 0) + 1
            dia = (c.banho_fim - timedelta(hours=3)).strftime('%d/%m')
            por_dia[dia] = por_dia.get(dia, 0) + 1
            dd = c.to_dict()
            peso_por_dia[dia] = round(peso_por_dia.get(dia, 0) + dd['peso_total'], 2)
            area_por_dia[dia] = round(area_por_dia.get(dia, 0) + dd['area_total'], 3)
            peso_total_geral += dd['peso_total']
            area_total_geral += dd['area_total']
            pecas_total_geral += dd['qtd_total']
        return {
            'total': len(cards), 'normais': normais, 'retrabalhos': retrab,
            'em_andamento': len(ativos),
            'banho_normal': banho_normal, 'banho_retrabalho': banho_retrab,
            'media_prep': round(sum(tp) / len(tp), 1) if tp else 0,
            'media_banho': round(sum(tb) / len(tb), 1) if tb else 0,
            'media_espera': round(sum(esperas) / len(esperas), 1) if esperas else 0,
            'por_processo': por_proc, 'por_dia': por_dia,
            'peso_por_dia': peso_por_dia, 'area_por_dia': area_por_dia,
            'peso_total_geral': round(peso_total_geral, 1),
            'area_total_geral': round(area_total_geral, 2),
            'pecas_total_geral': pecas_total_geral,
            'ativos': [c.to_dict() for c in ativos],
            'registros': [c.to_dict() for c in sorted(cards, key=lambda x: x.id, reverse=True)[:200]],
        }
    finally:
        db.close()


def _parse_datas():
    def pd(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None
    return pd(request.args.get('de')), pd(request.args.get('ate'))


@app.route('/api/dashboard/dados')
@login_required('admin')
def api_dashboard_dados():
    de, ate = _parse_datas()
    return jsonify(_coletar_dados(de, ate))


@app.route('/api/painel/dados')
def api_painel_dados():
    de, ate = _parse_datas()
    return jsonify(_coletar_dados(de, ate))


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────────────────
def _estilo_cabecalho(ws, headers):
    fill = PatternFill("solid", fgColor="0F3D5C")
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color='D0D7DE')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 28


def _gerar_excel(tipo):
    db = Session()
    try:
        cards = db.query(Card).filter_by(estado=ST_CONCLUIDO).order_by(Card.id).all()
        wb = Workbook()
        ws = wb.active

        if tipo == 'prebanho':
            ws.title = 'Pre-Banho'
            headers = ['ID', 'Cesto', 'OP (Ordem)', 'Código', 'Texto breve', 'Qtd',
                       'Área total (m²)', 'Peso total (kg)',
                       'Processo', 'Tipo', 'Operador Prep', 'Nº oper.',
                       'Início Prep - Data', 'Início Prep - Hora',
                       'Fim Prep - Data', 'Fim Prep - Hora',
                       'Tempo prep (min)', 'Tempo parada (min)',
                       'Pausas (por motivo)', 'Observação']
            larg = [6, 7, 14, 14, 30, 7, 14, 14, 22, 12, 18, 9, 16, 14, 16, 14, 13, 14, 30, 28]

        elif tipo == 'banho':
            ws.title = 'Banho'
            headers = ['ID', 'Cesto', 'OP (Ordem)', 'Código', 'Texto breve', 'Qtd',
                       'Área total (m²)', 'Peso total (kg)',
                       'Processo', 'Tipo',
                       'Operador Banho Início', 'Operador Banho Fim',
                       'Fim Prep - Data', 'Fim Prep - Hora',
                       'Início Banho - Data', 'Início Banho - Hora',
                       'Final Banho - Data', 'Final Banho - Hora',
                       'Tempo espera (min)', 'Tempo banho (min)', 'Observação banho']
            larg = [6, 7, 14, 14, 30, 7, 14, 14, 22, 12, 20, 20, 16, 14, 16, 14, 16, 14, 14, 14, 28]

        else:  # geral
            ws.title = 'Geral'
            headers = ['ID', 'Cesto', 'OP (Ordem)', 'Código', 'Texto breve', 'Qtd',
                       'Área total (m²)', 'Peso total (kg)',
                       'Processo', 'Tipo',
                       'Operador Prep', 'Nº oper.',
                       'Início Prep - Data', 'Início Prep - Hora',
                       'Fim Prep - Data', 'Fim Prep - Hora',
                       'Tempo prep (min)', 'Tempo parada (min)',
                       'Operador Banho Início', 'Operador Banho Fim',
                       'Início Banho - Data', 'Início Banho - Hora',
                       'Final Banho - Data', 'Final Banho - Hora',
                       'Tempo espera (min)', 'Tempo banho (min)', 'Total prep+banho (min)',
                       'Observação Prep', 'Observação Banho']
            larg = [6, 7, 14, 14, 30, 7, 14, 14, 22, 12, 18, 9, 16, 14, 16, 14, 13, 14,
                    20, 20, 16, 14, 16, 14, 14, 14, 16, 28, 28]

        _estilo_cabecalho(ws, headers)

        for c in cards:
            dd = c.to_dict()
            itens = dd['itens'] or [{'ordem': dd['ordem'], 'material': dd['material'],
                                     'texto_breve': dd['texto_breve'], 'quantidade': dd['quantidade']}]
            total_prep_banho = round((dd['prep_minutos'] or 0) + (dd['banho_minutos'] or 0), 1)

            for it in itens:
                a_unit, p_unit = _area_peso_do_codigo(it.get('material', ''))
                q_it = int(it.get('quantidade') or 0)
                area_it = round(a_unit * q_it, 3)
                peso_it = round(p_unit * q_it, 2)
                if tipo == 'prebanho':
                    ws.append([
                        dd['id'], dd['numero_cesto'], it['ordem'], it['material'],
                        it['texto_breve'], it['quantidade'], area_it, peso_it,
                        dd['processo'], dd['tipo'],
                        dd['operador_prep'], dd['n_operadores'],
                        dd['prep_inicio_data'], dd['prep_inicio_hora'],
                        dd['prep_fim_data'], dd['prep_fim_hora'],
                        dd['prep_minutos'], dd['total_pausa_min'],
                        dd['pausas']['texto'], dd['observacao']
                    ])
                elif tipo == 'banho':
                    ws.append([
                        dd['id'], dd['numero_cesto'], it['ordem'], it['material'],
                        it['texto_breve'], it['quantidade'], area_it, peso_it,
                        dd['processo'], dd['tipo'],
                        dd['operador_banho_inicio'], dd['operador_banho_fim'],
                        dd['prep_fim_data'], dd['prep_fim_hora'],
                        dd['banho_inicio_data'], dd['banho_inicio_hora'],
                        dd['banho_fim_data'], dd['banho_fim_hora'],
                        dd['espera_min'], dd['banho_minutos'], dd['obs_banho']
                    ])
                else:  # geral
                    ws.append([
                        dd['id'], dd['numero_cesto'], it['ordem'], it['material'],
                        it['texto_breve'], it['quantidade'], area_it, peso_it,
                        dd['processo'], dd['tipo'],
                        dd['operador_prep'], dd['n_operadores'],
                        dd['prep_inicio_data'], dd['prep_inicio_hora'],
                        dd['prep_fim_data'], dd['prep_fim_hora'],
                        dd['prep_minutos'], dd['total_pausa_min'],
                        dd['operador_banho_inicio'], dd['operador_banho_fim'],
                        dd['banho_inicio_data'], dd['banho_inicio_hora'],
                        dd['banho_fim_data'], dd['banho_fim_hora'],
                        dd['espera_min'], dd['banho_minutos'], total_prep_banho,
                        dd['observacao'], dd['obs_banho']
                    ])

        for i, w in enumerate(larg, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = 'A2'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    finally:
        db.close()


@app.route('/api/download/prebanho')
@login_required('admin')
def download_prebanho():
    buf = _gerar_excel('prebanho')
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'prebanho_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/download/banho')
@login_required('admin')
def download_banho():
    buf = _gerar_excel('banho')
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'banho_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/download/geral')
@login_required('admin')
def download_geral():
    buf = _gerar_excel('geral')
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'relatorio_geral_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.teardown_appcontext
def remove_session(exc=None):
    Session.remove()


init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
